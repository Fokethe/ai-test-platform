#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试用例模板生成脚本
根据规范生成4种测试用例Excel模板
"""

import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter


def create_functional_template():
    """创建功能测试用例模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "功能测试用例"
    
    # 表头
    headers = [
        "用例ID", "所属模块", "用例标题", "用例类型", "优先级",
        "前置条件", "测试步骤", "预期结果", "测试数据", "备注"
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 列宽设置
    column_widths = [12, 15, 30, 12, 10, 25, 40, 40, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 示例数据行（展示格式）
    example_data = [
        ["TC-FUN-001", "用户管理", "管理员登录系统成功", "正向", "P0",
         "1. 系统已部署\n2. 管理员账号已创建", 
         "1. 打开登录页面\n2. 输入正确账号密码\n3. 点击登录按钮",
         "1. 登录成功\n2. 跳转首页\n3. 显示管理员信息", 
         "账号：admin\n密码：Admin123", "需验证权限加载"]
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # 用例ID列颜色（示例：浅蓝）
            if col_idx == 1:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # 优先级列颜色
            if col_idx == 5:
                if value == "P0":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif value == "P1":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    return wb


def create_api_template():
    """创建接口测试用例模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "接口测试用例"
    
    headers = [
        "用例ID", "接口名称", "请求方法", "接口路径", "用例标题",
        "用例类型", "优先级", "前置条件", "请求头", "请求参数",
        "预期状态码", "预期响应", "数据库断言", "备注"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    column_widths = [12, 15, 12, 25, 30, 12, 10, 25, 25, 30, 12, 30, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.freeze_panes = "A2"
    
    # 示例数据
    example_data = [
        ["TC-API-001", "用户登录", "POST", "/api/v1/auth/login", "正确账号密码登录成功",
         "正向", "P0", "用户已注册，账号状态正常",
         "Content-Type: application/json",
         '{"username":"testuser","password":"Test123456"}',
         "200", '{"code":0,"token":"xxx","userId":"123"}',
         "更新last_login_time", "需验证Token有效期"]
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            if col_idx == 1:
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            
            if col_idx == 7:
                if value == "P0":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif value == "P1":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    return wb


def create_performance_template():
    """创建性能测试用例模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "性能测试用例"
    
    headers = [
        "用例ID", "用例标题", "测试类型", "测试目的", "测试范围",
        "负载模型", "测试数据", "预期TPS", "预期响应时间", "错误率阈值",
        "资源阈值", "通过标准", "环境要求", "备注"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    column_widths = [12, 30, 12, 30, 25, 35, 25, 12, 20, 12, 20, 30, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.freeze_panes = "A2"
    
    # 示例数据
    example_data = [
        ["TC-PERF-001", "用户下单接口负载测试", "负载测试",
         "验证下单接口在高峰期的性能表现", "POST /api/v1/orders",
         "并发用户：50→100→200→500，每阶段5分钟",
         "10000个用户账号，商品库存充足",
         "> 100", "平均<200ms, P95<500ms", "< 0.1%",
         "CPU<80%, 内存<80%",
         "500并发下响应时间<1s，无错误",
         "8核16G服务器，千兆网络",
         "需要监控数据库连接池"]
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            if col_idx == 1:
                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    return wb


def create_automation_template():
    """创建自动化测试用例模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "自动化测试用例"
    
    headers = [
        "用例ID", "用例标题", "业务价值", "技术可行性", "维护成本",
        "实现优先级", "建议分层", "建议框架", "实现难度", "预估工时",
        "依赖准备", "关键步骤", "断言点", "数据策略", "备注"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    column_widths = [15, 30, 10, 10, 10, 10, 10, 15, 10, 10, 25, 30, 25, 25, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.freeze_panes = "A2"
    
    # 示例数据
    example_data = [
        ["TC-AUTO-001 (TC-API-001)", "用户登录接口自动化测试",
         "高", "高", "低", "P0", "接口测试", "Pytest + Requests",
         "简单", "2小时", "测试环境已部署，测试账号已准备",
         "1.构造请求参数\n2.调用接口\n3.验证响应",
         "状态码200、返回Token、Token格式正确",
         "使用固定测试账号，无需清理",
         "建议加入CI/CD流程"]
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            if col_idx == 1:
                cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
            
            if col_idx == 6:
                if value == "P0":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif value == "P1":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    return wb


def main():
    """主函数：生成所有模板"""
    import os
    
    # 获取assets目录路径
    script_dir = os.path.dirname(os.path.abspath(__file__))
    assets_dir = os.path.join(os.path.dirname(script_dir), "assets")
    
    templates = {
        "功能测试用例模板.xlsx": create_functional_template,
        "接口测试用例模板.xlsx": create_api_template,
        "性能测试用例模板.xlsx": create_performance_template,
        "自动化测试用例模板.xlsx": create_automation_template,
    }
    
    print("开始生成测试用例模板...")
    print("-" * 50)
    
    for filename, create_func in templates.items():
        filepath = os.path.join(assets_dir, filename)
        try:
            wb = create_func()
            wb.save(filepath)
            print(f"✅ 已生成: {filename}")
        except Exception as e:
            print(f"❌ 生成失败: {filename}")
            print(f"   错误: {e}")
    
    print("-" * 50)
    print("模板生成完成！")
    print(f"保存位置: {assets_dir}")


if __name__ == "__main__":
    main()
