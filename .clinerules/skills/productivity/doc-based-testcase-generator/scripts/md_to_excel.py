#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将Markdown格式的测试用例转换为Excel文件
"""

import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

def parse_test_cases(md_content):
    """解析Markdown内容，提取测试用例"""
    test_cases = []
    
    # 按表格分割内容
    tables = re.findall(r'\|[^|]+\|[^|]+\|[^|]+\|[^|]+\|[^|]+\|[^|]+\|[^|]+\|[^|]+\|[^|]+\|[^|]+\|[^|]+\|', md_content)
    
    current_module = ""
    current_sub_module = ""
    
    lines = md_content.split('\n')
    
    for i, line in enumerate(lines):
        # 识别模块标题
        if line.startswith('## '):
            current_module = line.replace('## ', '').strip()
            continue
        if line.startswith('### '):
            current_sub_module = line.replace('### ', '').strip()
            continue
            
        # 解析表格行
        if line.startswith('| TC-'):
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            if len(cells) >= 8:
                test_case = {
                    '用例ID': cells[0],
                    '需求模块/功能点': cells[1],
                    '用例标题': cells[2],
                    '用例类型': cells[3],
                    '优先级': cells[4],
                    '前置条件': cells[5].replace('<br>', '\n'),
                    '测试步骤': cells[6].replace('<br>', '\n'),
                    '预期结果': cells[7].replace('<br>', '\n')
                }
                test_cases.append(test_case)
    
    return test_cases

def create_excel(test_cases, output_file):
    """创建Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "功能测试用例"
    
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='微软雅黑', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 标题行
    headers = ['用例ID', '需求模块/功能点', '用例标题', '用例类型', '优先级', 
               '前置条件', '测试步骤', '预期结果']
    
    # 写入标题
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row_idx, tc in enumerate(test_cases, 2):
        ws.cell(row=row_idx, column=1, value=tc['用例ID'])
        ws.cell(row=row_idx, column=2, value=tc['需求模块/功能点'])
        ws.cell(row=row_idx, column=3, value=tc['用例标题'])
        ws.cell(row=row_idx, column=4, value=tc['用例类型'])
        ws.cell(row=row_idx, column=5, value=tc['优先级'])
        ws.cell(row=row_idx, column=6, value=tc['前置条件'])
        ws.cell(row=row_idx, column=7, value=tc['测试步骤'])
        ws.cell(row=row_idx, column=8, value=tc['预期结果'])
        
        # 应用样式
        for col in range(1, 9):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # 优先级颜色标记
            if col == 5:  # 优先级列
                if tc['优先级'] == 'P0':
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                elif tc['优先级'] == 'P1':
                    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    # 设置列宽
    column_widths = {
        'A': 12,   # 用例ID
        'B': 20,   # 需求模块/功能点
        'C': 35,   # 用例标题
        'D': 10,   # 用例类型
        'E': 8,    # 优先级
        'F': 30,   # 前置条件
        'G': 40,   # 测试步骤
        'H': 35    # 预期结果
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 30
    for row in range(2, len(test_cases) + 2):
        ws.row_dimensions[row].height = 60
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(output_file)
    print(f"Excel文件已生成: {output_file}")
    print(f"共 {len(test_cases)} 条测试用例")

def main():
    md_file = r"f:\test\skills\productivity\doc-based-testcase-generator\testcases\钢结构检测管理系统_功能测试用例.md"
    output_file = r"f:\test\skills\productivity\doc-based-testcase-generator\testcases\钢结构检测管理系统_功能测试用例.xlsx"
    
    if not os.path.exists(md_file):
        print(f"错误: 找不到文件 {md_file}")
        return
    
    with open(md_file, 'r', encoding='utf-8') as f:
        md_content = f.read()
    
    test_cases = parse_test_cases(md_content)
    
    if not test_cases:
        print("未找到测试用例数据")
        return
    
    create_excel(test_cases, output_file)

if __name__ == "__main__":
    main()
